import re

import openpyxl
from django.contrib import messages
from django.shortcuts import render
from transliterate import translit
from data.models import Data


class UploadingData(object):
    """Загрузка данных xlsx-файла в базу данных."""

    model = Data

    def __init__(self, data):
        """Получает файл из браузера."""
        data = data
        self.uploaded_file = data.get('file')
        self.parsing()

    def get_headers(self):
        """Определяет названия заголовков таблицы."""
        sheet = self.sheet
        headers = {}
        for column in range(1, sheet.max_column+1):
            ru_value = sheet.cell(1, column).value
            en_value = translit(         # Если заголовки полей на русском, переводит на английский
                ru_value, reversed=True  # чтобы была возможность создать поля модели.
                ).replace('  ', '_').lower()
            value = re.sub(r"['-/.]", r"", en_value)
            headers[column] = value.replace(' ', '_')
        return headers

    def parsing(self):
        """Переносит данные из xlsx-файла в базу данных."""
        book = openpyxl.open(self.uploaded_file, read_only=True)
        sheet = book.active
        self.sheet = sheet
        headers = self.get_headers()
        print(headers)

        for row in range(2, sheet.max_row+1):
            flag = ''
            dict_data = {}
            for column in range(1, sheet.max_column+1):
                if headers[column] is None:
                    continue
                value_column_1 = sheet.cell(row, 1).value
                if value_column_1 is None:
                    flag = 'stop'
                value = sheet.cell(row, column).value
                field_name = headers[column]
                dict_data[field_name] = value
            if flag == 'stop':
                break
            Data.objects.create(**dict_data)


def download_data(request):
    """View-функция загрузки файла со страницы в браузере."""
    if request.POST:
        file = request.FILES['file']
        uploading_file = UploadingData({'file': file})
        if uploading_file:
            messages.success(request, 'Успешная загрузка')
        else:
            messages.error(request, 'Ошибка при загрузке')
    return render(request, 'addpage.html', {'title': 'Добавте файл'})
